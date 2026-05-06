#!/usr/bin/env python3
# Python 3.9 compatible -- no PEP 604 unions, no match/case, no list[...] generics.
"""HKEX A1 scraper for /a1-pipeline-update Phase 0.

Hard contract: the HKEX Application Proof JSON feed at
  https://www1.hkexnews.hk/ncms/json/eds/app_{YYYY}_sehk_{e|c}.json
If that URL pattern changes, this module breaks -- catch JSONDecodeError and
KeyError in the caller and surface a clear message per edge case #4 in the plan.
"""

import datetime
import difflib
import os
import re
import shutil
import time
from typing import Any, Callable, Dict, List, Optional, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

# ============================================================================
# Module constants
# ============================================================================

HKEX_JSON_BASE = "https://www1.hkexnews.hk/ncms/json/eds/"
HKEX_DOC_BASE = "https://www1.hkexnews.hk/app/"
USER_AGENT = "Mozilla/5.0 (A1-Pipeline-Update/1.0)"
REQ_TIMEOUT = 30
REQ_RETRY = 3
REQ_BACKOFF = 2.0   # seconds, doubled each retry
REQ_DELAY = 0.8     # minimum gap between successive HKEX requests

EN_KEYWORDS = (
    "pharma", "bio", "therap", "medic", "medi", "health",
    "diagn", "medtech", "cdmo", "vaccine", "drug", "gene",
    "oncolog", "clinic", "hospital", "surger", "life scien",
    "genomic", "immun", "dental",
)
CN_KEYWORDS = (
    "\u836f", "\u91ab", "\u533b", "\u5065", "\u8bca", "\u8a3a",
    "\u7642", "\u7597", "\u751f\u7269", "\u5236\u836f", "\u88fd\u85e5",
    "\u533b\u7597", "\u91ab\u7642", "\u533b\u836f", "\u91ab\u85e5",
    "\u5065\u5eb7", "\u75ab\u82d7", "\u57fa\u56e0",
    "\u8bca\u65ad", "\u8a3a\u65b7",
)
SUFFIX_B_P = ("- B", "-B", "- P", "-P")

CANONICAL_SECTORS = (
    "Pharma / Biotech", "MedTech",
    "Services (TCM / CXO / Hospital / PBM)",
    "Diagnostics / Tools", "CDMO", "Healthcare Tech",
    "Consumer Health", "Vaccine", "Distributor",
)

# Firecrawl narrative extraction schema + prompt for cols F, G, H, M.
# Col I (sponsor) stays on regex because sponsor-bank text is templated and
# regex is both deterministic and reliable. Col J/K/L stay on pdfplumber
# because numbers need deterministic table parsing, not LLM inference.
FIRECRAWL_NARRATIVE_SCHEMA = {
    "type": "object",
    "properties": {
        "shareholder_structure": {
            "type": ["string", "null"],
            "enum": [
                "H-share", "Red Chip", "VIE", None,
            ],
            "description": (
                "Corporate structure of the issuer. Three values only: "
                "'H-share' for PRC-domiciled joint stock companies "
                "(\u80a1\u4efd\u6709\u9650\u516c\u53f8 incorporated in PRC); "
                "'Red Chip' for ANY offshore-holdco structure "
                "(Cayman / BVI / Bermuda) WITHOUT a VIE in the chain; "
                "'VIE' when a variable interest entity is in the chain. "
                "Do NOT return 'Cayman holdco' or 'BVI holdco' \u2014 "
                "those collapse into 'Red Chip'. The specific incorporation "
                "jurisdiction (Cayman vs BVI vs Bermuda) is captured in col N "
                "(Highlights) rather than col F."
            ),
        },
        "business_model": {
            "type": ["string", "null"],
            "description": (
                "Format '<stage>; <one-line description of business model or "
                "core asset status>' where stage is exactly one of "
                "'Commercial-stage' | 'Clinical-stage' | 'Pre-clinical' | "
                "'Commercialized'. Example: 'Clinical-stage; 5 assets in "
                "Phase II/III oncology pipeline, pre-revenue'."
            ),
        },
        "sector": {
            "type": ["string", "null"],
            "enum": list(CANONICAL_SECTORS) + [None],
            "description": (
                "Canonical sector classification. Choose based on the main "
                "business described in the prospectus, not on incidental "
                "vendor/supplier references."
            ),
        },
        "lead_asset": {
            "type": ["string", "null"],
            "description": (
                "One-line description of the core product or lead pipeline "
                "asset, including indication and development stage when "
                "available. Example: 'SRSD107, siRNA for HBV, Phase I'."
            ),
        },
    },
    "required": [
        "shareholder_structure", "business_model", "sector", "lead_asset",
    ],
}

FIRECRAWL_NARRATIVE_PROMPT = (
    "Extract structured fields from this HKEX prospectus chapter. "
    "Return null for any field not clearly stated in the text. "
    "Do not guess. For shareholder_structure, read the Corporate "
    "Structure / History and Reorganization section carefully and pick "
    "EXACTLY ONE of three values: 'H-share' (PRC-domiciled 股份有限公司), "
    "'Red Chip' (offshore Cayman / BVI / Bermuda holdco WITHOUT VIE — the "
    "only thing you need to detect is whether a variable interest entity / "
    "可变利益实体 / contractual arrangement controls a PRC OpCo), or 'VIE' "
    "(a variable interest entity IS in the chain). The narrower 'Cayman "
    "holdco' / 'BVI holdco' labels are NOT valid outputs — those collapse "
    "into 'Red Chip'. For sector, match the issuer's MAIN business, "
    "ignoring references to third-party suppliers (e.g., a biotech using "
    "a CDMO is still 'Pharma / Biotech', not 'CDMO')."
)

# F+G robustness tags. One tag per NEW/REFRESH row, surfaced in Phase 4
# (Gate 2) and aggregated by the Phase 4 robustness counter. Borrowed from
# /web-research Tag Vocabulary (Rule 4 + Rule 10) -- see
# `.claude/skills/a1-pipeline-update/SKILL.md` Phase 4 Robustness Tag
# Vocabulary section for the full decision table.
FG_ROBUSTNESS_TAGS = (
    "verified_fg_prospectus",        # Phase 0: prospectus + high confidence
    "single_source_prospectus_fg",   # Phase 0: prospectus extracted but conf != "high"
    "web_cross_checked_fg",          # Phase 3: prospectus failed -> >=2 independent web sources
    "single_source_family_fg",       # Phase 3: prospectus failed -> only one source family
    "conflicting_fg",                # Prospectus disagrees with secondary; needs user resolve
    "not_found_fg",                  # All Tier-1 attempts failed; needs user accept_blank
)

FG_ROBUSTNESS_COL_N_SUFFIX = {
    "verified_fg_prospectus":      "[VERIFIED F+G — prospectus]",
    "single_source_prospectus_fg": "[Single source — prospectus only]",
    "web_cross_checked_fg":        "[Web cross-checked F+G]",
    "single_source_family_fg":     "[Single source family — verify]",
    "conflicting_fg":               "[Conflicting — used prospectus]",
    "not_found_fg":                 "[NOT FOUND — searched: prospectus, firecrawl_search top-3, IR]",
}

# Regex matching ANY robustness suffix at the end of col N (covers customized
# suffixes too -- e.g., a not_found_fg row whose searched-list got tweaked,
# or a conflicting_fg row whose discarded-source footnote was appended).
# Used by apply_fg_robustness_tag to strip the existing suffix before adding
# a new one (so re-classification doesn't double-stamp).
_FG_ROBUSTNESS_SUFFIX_RE = re.compile(
    r"\s*\[(?:VERIFIED F\+G|Single source|Web cross-checked|Conflicting|NOT FOUND)"
    r"[^\]]*\]\s*$"
)

LEGAL_SUFFIX_STRIP = (
    "co., ltd.", "co.,ltd.", "co. ltd.",
    "company limited", "limited", "ltd.", "ltd",
    "inc.", "incorporated", "holdings", "group",
    "corporation", "corp.", "corp",
)

CHAPTER_VARIANTS = {
    "summary": [
        "summary", "executive summary", "overview and summary",
        "\u6982\u8981", "\u6458\u8981", "\u6982\u89bd",
    ],
    "business": [
        "business", "our business", "business overview",
        "\u696d\u52d9", "\u4e1a\u52a1", "\u4e1a\u52a1\u6982\u89c8",
        "\u6211\u5011\u7684\u696d\u52d9", "\u6211\u4eec\u7684\u4e1a\u52a1",
    ],
    "financial": [
        "financial information",
        "summary of historical financial information",
        "summary historical financial information",
        "summary financial information",
        "accountants' report", "accountant's report",
        "\u8ca1\u52d9\u8cc7\u6599", "\u8d22\u52a1\u8d44\u6599",
        "\u5386\u53f2\u8d22\u52a1\u8d44\u6599", "\u6b77\u53f2\u8ca1\u52d9\u8cc7\u6599",
        "\u6703\u8a08\u5e2b\u5831\u544a", "\u4f1a\u8ba1\u5e08\u62a5\u544a",
    ],
    # PARTIES chapter: holds Sole Sponsor / Joint Sponsor info that the regex
    # extractor needs for col I. Variants cover both "DIRECTORS AND PARTIES
    # INVOLVED" (most filings) and "DIRECTORS, SUPERVISORS AND PARTIES INVOLVED"
    # (A-share dual-listings with PRC supervisor board).
    "parties": [
        "directors and parties involved",
        "directors, supervisors and parties involved",
        "directors supervisors and parties involved",
        "parties involved",
        "\u8463\u4e8b\u53ca\u53c3\u8207\u4eba\u58eb",
        "\u8463\u4e8b\u3001\u76e3\u4e8b\u53ca\u53c3\u8207\u4eba\u58eb",
        "\u8463\u4e8b\u3001\u9ad8\u7ba1\u53c3\u8207",
    ],
    # HISTORY chapter: Corporate Structure section used by the LLM to confirm
    # F (H-share / Red Chip / VIE) when SUMMARY alone is ambiguous.
    "history": [
        "history, development and corporate structure",
        "history development and corporate structure",
        "history and corporate structure",
        "\u6b77\u53f2\u3001\u767c\u5c55\u53ca\u516c\u53f8\u67b6\u69cb",
        "\u5386\u53f2\u53d1\u5c55\u4e0e\u516c\u53f8\u67b6\u6784",
    ],
}

# Module-global state
_last_request_ts = 0.0
warnings_log: List[str] = []


# ============================================================================
# HTTP helper with rate limit + retry
# ============================================================================

class HkexFetchError(RuntimeError):
    """Raised when HKEX fetch permanently fails after retries."""
    pass


def _http_get(url: str, as_json: bool = False) -> Any:
    """GET with rate limit + exponential-backoff retry. Returns Response or dict."""
    global _last_request_ts
    last_err = None
    for attempt in range(REQ_RETRY):
        elapsed = time.time() - _last_request_ts
        if elapsed < REQ_DELAY:
            time.sleep(REQ_DELAY - elapsed)
        try:
            r = requests.get(
                url,
                headers={"User-Agent": USER_AGENT},
                timeout=REQ_TIMEOUT,
            )
            _last_request_ts = time.time()
            if r.status_code == 200:
                if as_json:
                    return r.json()
                return r
            if r.status_code in (429, 500, 502, 503, 504):
                last_err = "HTTP {}".format(r.status_code)
                time.sleep(REQ_BACKOFF * (2 ** attempt))
                continue
            r.raise_for_status()
        except (requests.ConnectionError, requests.Timeout) as e:
            last_err = str(e)
            if attempt == REQ_RETRY - 1:
                raise HkexFetchError(
                    "fetch failed after {} retries: {} -- {}".format(
                        REQ_RETRY, url, last_err))
            time.sleep(REQ_BACKOFF * (2 ** attempt))
    raise HkexFetchError(
        "fetch failed after {} retries: {} -- {}".format(
            REQ_RETRY, url, last_err))


# ============================================================================
# Feed fetching
# ============================================================================

def fetch_hkex_feed(
    year: int,
    board: str = "sehk",
    lang: str = "en",
    active_only: bool = False,
) -> Dict[str, Any]:
    """Fetch and parse the HKEX Application Proof JSON feed."""
    if board not in ("sehk", "gem"):
        raise ValueError("board must be 'sehk' or 'gem'")
    if lang not in ("en", "c"):
        raise ValueError("lang must be 'en' or 'c'")
    lang_code = "e" if lang == "en" else "c"
    if active_only:
        fname = "appactive_app_{board}_{lc}.json".format(board=board, lc=lang_code)
    else:
        fname = "app_{y}_{board}_{lc}.json".format(y=year, board=board, lc=lang_code)
    url = HKEX_JSON_BASE + fname
    data = _http_get(url, as_json=True)
    if not isinstance(data, dict) or "app" not in data:
        raise HkexFetchError(
            "unexpected feed shape at {}: keys={}".format(
                url,
                list(data.keys()) if isinstance(data, dict) else type(data).__name__))
    return data


def _parse_hkex_date(s: Any) -> Optional[datetime.date]:
    """Parse dd/mm/yyyy string into date; tolerate None/empty/non-string."""
    if not s or not isinstance(s, str):
        return None
    try:
        return datetime.datetime.strptime(s.strip(), "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None


def _pick_latest_ap(ls: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """From a record's `ls` array, pick the Application Proof we care about.

    Preference: has Multi-Files TOC (u2), then most-recent submission date.
    """
    ap = [
        s for s in ls
        if isinstance(s.get("nF"), str)
        and s["nF"].startswith("Application Proof")
    ]
    if not ap:
        return None
    ap.sort(
        key=lambda s: (
            bool(s.get("u2")),
            _parse_hkex_date(s.get("d", "")) or datetime.date.min,
        ),
        reverse=True,
    )
    return ap[0]


def fetch_lifesci_candidates(
    year: int,
    since: Optional[datetime.date] = None,
    include_gem: bool = False,
) -> List[Dict[str, Any]]:
    """Return unified EN+CN candidate records with absolute TOC URLs resolved."""
    boards = ["sehk"]
    if include_gem:
        boards.append("gem")
    candidates: List[Dict[str, Any]] = []
    for board in boards:
        try:
            feed_en = fetch_hkex_feed(year, board=board, lang="en")
            feed_cn = fetch_hkex_feed(year, board=board, lang="c")
        except HkexFetchError as e:
            warnings_log.append("feed fetch failed for board={}: {}".format(board, e))
            continue
        cn_by_id = {rec.get("id"): rec for rec in feed_cn.get("app", [])}
        for rec in feed_en.get("app", []):
            ap = _pick_latest_ap(rec.get("ls", []))
            if not ap:
                continue
            filing_date = _parse_hkex_date(ap.get("d") or rec.get("d", ""))
            if filing_date is None:
                continue
            if since is not None and filing_date < since:
                continue
            cn_rec = cn_by_id.get(rec.get("id"), {})
            name_cn = cn_rec.get("a") if isinstance(cn_rec, dict) else None
            name_en = (rec.get("a") or "").strip()
            if isinstance(name_cn, str) and name_cn.strip() == name_en:
                name_cn = None
            u1 = ap.get("u1")
            u2 = ap.get("u2")
            # HKEX uses "#" or empty string for lapsed/withdrawn/unavailable PDFs.
            if u1 in (None, "", "#"):
                u1 = None
            if u2 in (None, "", "#"):
                u2 = None
            full_pdf_url = urljoin(HKEX_DOC_BASE, u1) if u1 else None
            multi_files_url = urljoin(HKEX_DOC_BASE, u2) if u2 else None
            # Skip records where both URLs are unavailable -- we have no way
            # to fetch the prospectus content.
            if full_pdf_url is None and multi_files_url is None:
                continue
            candidates.append({
                "id": rec.get("id"),
                "board": board,
                "filing_date": filing_date,
                "name_en": name_en,
                "name_cn": name_cn.strip() if isinstance(name_cn, str) else None,
                "filing_type": ap.get("nF", ""),
                "full_pdf_url": full_pdf_url,
                "multi_files_url": multi_files_url,
                "raw": rec,
            })
    candidates.sort(key=lambda c: c["filing_date"], reverse=True)
    return candidates


# ============================================================================
# Keyword pre-filter
# ============================================================================

def is_lifesci_candidate(name_en: str, name_cn: Optional[str]) -> bool:
    """EN keyword OR CN keyword OR -B/-P suffix match."""
    en = (name_en or "").lower()
    if any(kw in en for kw in EN_KEYWORDS):
        return True
    stripped = en.rstrip()
    if any(stripped.endswith(sfx.lower()) for sfx in SUFFIX_B_P):
        return True
    if name_cn and any(kw in name_cn for kw in CN_KEYWORDS):
        return True
    return False


def _match_reasons(name_en: str, name_cn: Optional[str]) -> List[str]:
    """Return the list of matching reasons for a record."""
    reasons: List[str] = []
    en = (name_en or "").lower()
    if any(kw in en for kw in EN_KEYWORDS):
        reasons.append("EN kw")
    stripped = en.rstrip()
    if any(stripped.endswith(sfx.lower()) for sfx in SUFFIX_B_P):
        reasons.append("suffix")
    if name_cn and any(kw in name_cn for kw in CN_KEYWORDS):
        reasons.append("CN kw")
    return reasons


def filter_candidates(
    candidates: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Partition (passed, dropped); annotate each passed record with reason."""
    passed: List[Dict[str, Any]] = []
    dropped: List[Dict[str, Any]] = []
    for c in candidates:
        reasons = _match_reasons(c.get("name_en", ""), c.get("name_cn"))
        if reasons:
            enriched = dict(c)
            enriched["reason"] = "+".join(reasons)
            passed.append(enriched)
        else:
            dropped.append(c)
    return passed, dropped


# ============================================================================
# Master tracker load + name normalization + dedup classification
# ============================================================================

def normalize_company_name(name: str) -> str:
    """Canonical form for name-matching across feed and master file.

    Lowercases, strips -B/-P/dagger suffixes, removes parenthetical
    "(formerly known as ...)" text, strips common legal suffixes and
    collapses whitespace, so that 'Sirius Therapeutics - B' and
    'Sirius Therapeutics' collapse to the same key.
    """
    if not name:
        return ""
    s = str(name).strip()
    s = s.replace("\u2020", "")  # dagger -- lapsed marker
    s = s.lower().strip()
    for sfx in ("- b", "-b", "- p", "-p"):
        if s.endswith(sfx):
            s = s[: -len(sfx)].rstrip()
            break
    s = re.sub(r"\(formerly[^)]*\)", "", s).strip()
    s = re.sub(r"\(previously[^)]*\)", "", s).strip()
    changed = True
    while changed:
        changed = False
        for suffix in LEGAL_SUFFIX_STRIP:
            if s.endswith(suffix):
                s = s[: -len(suffix)].rstrip(",. ")
                changed = True
                break
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip(",. ")
    return s


def _parse_master_date(value: Any) -> Optional[datetime.date]:
    """Coerce a master col C cell value to a datetime.date.

    Handles the three observed shapes: datetime objects (rows 3-6),
    'dd/mm/yyyy' strings with possible trailing whitespace (rows 92+),
    and rare Excel serial numerics.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        s = value.strip().rstrip(" .,")
        parsed = _parse_hkex_date(s)
        if parsed is not None:
            return parsed
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            d = from_excel(value)
            if isinstance(d, datetime.datetime):
                return d.date()
            if isinstance(d, datetime.date):
                return d
        except Exception:
            pass
    return None


def load_master_tracker(xlsx_path: str) -> Dict[str, Any]:
    """Read-only snapshot of the current master tracker.

    Returns a dict with path/sheet_title/header_row/data_start_row/a1_metadata
    and `rows` -- a list of per-row dicts with parsed filing_date, normalized
    name, CN name, sector, J/K/L/N values, and a dagger flag.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    a1_text = ws["A1"].value or ""
    header_row = 2
    data_start_row = 3
    rows: List[Dict[str, Any]] = []
    for r in range(data_start_row, ws.max_row + 1):
        name_en_raw = ws.cell(row=r, column=4).value
        if not name_en_raw:
            continue
        name_cn_raw = ws.cell(row=r, column=5).value
        filing_raw = ws.cell(row=r, column=3).value
        sector = ws.cell(row=r, column=8).value
        j = ws.cell(row=r, column=10).value
        k = ws.cell(row=r, column=11).value
        l = ws.cell(row=r, column=12).value
        n = ws.cell(row=r, column=14).value
        filing_date = _parse_master_date(filing_raw)
        name_raw_str = str(name_en_raw).strip()
        rows.append({
            "row_idx": r,
            "name_raw": name_raw_str,
            "name_norm": normalize_company_name(name_raw_str),
            "name_cn": (str(name_cn_raw).strip() if name_cn_raw else None),
            "filing_date": filing_date,
            "filing_raw": filing_raw,
            "sector": sector,
            "j": j,
            "k": k,
            "l": l,
            "n": n,
            "has_dagger": "\u2020" in name_raw_str,
        })
    wb.close()
    return {
        "path": xlsx_path,
        "sheet_title": ws.title,
        "header_row": header_row,
        "data_start_row": data_start_row,
        "a1_metadata": a1_text,
        "rows": rows,
    }


def _find_master_match(
    cand: Dict[str, Any],
    master: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Return list of master rows matching candidate by EN norm or CN exact."""
    cand_norm = normalize_company_name(cand.get("name_en", ""))
    cand_cn = (cand.get("name_cn") or "").strip() or None
    matches: List[Dict[str, Any]] = []
    seen_rows = set()
    for row in master["rows"]:
        if cand_norm and row["name_norm"] == cand_norm:
            matches.append(row)
            seen_rows.add(row["row_idx"])
            continue
        if cand_cn and row.get("name_cn"):
            # Match on CN: normalize by stripping whitespace; accept Traditional/Simplified variants
            if cand_cn == row["name_cn"] or cand_cn in row["name_cn"] or row["name_cn"] in cand_cn:
                if row["row_idx"] not in seen_rows:
                    matches.append(row)
                    seen_rows.add(row["row_idx"])
    return matches


def classify_candidates(
    candidates: List[Dict[str, Any]],
    master: Dict[str, Any],
    stale_threshold_days: int = 90,
) -> Dict[str, List[Dict[str, Any]]]:
    """Bucket candidates into new / refresh / skip against the master tracker.

    Rules:
      NEW     : no name match in master
      REFRESH : name match AND candidate filing_date > master filing_date
                (fields_to_refresh = ['C']; + ['J','K','L','N'] if delta >= 90d)
      SKIP    : name match AND dates equal (or master newer, which is rare)
    """
    buckets: Dict[str, List[Dict[str, Any]]] = {
        "new": [], "refresh": [], "skip": [],
    }
    for cand in candidates:
        matches = _find_master_match(cand, master)
        enriched = dict(cand)
        enriched["qc_flags"] = list(cand.get("qc_flags", []))
        if not matches:
            enriched["bucket"] = "new"
            enriched["fields_to_refresh"] = [
                "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"
            ]
            buckets["new"].append(enriched)
            continue
        matches.sort(
            key=lambda m: m["filing_date"] or datetime.date.min,
            reverse=True,
        )
        best = matches[0]
        enriched["master_row_idx"] = best["row_idx"]
        enriched["master_filing_date"] = best["filing_date"]
        if len(matches) > 1:
            enriched["qc_flags"].append(
                "multi_row_match (rows={})".format(
                    ",".join(str(m["row_idx"]) for m in matches)
                )
            )
        cand_date = cand["filing_date"]
        master_date = best["filing_date"]
        if master_date is None:
            enriched["bucket"] = "refresh"
            enriched["fields_to_refresh"] = ["C", "J", "K", "L", "N"]
            enriched["date_delta_days"] = None
            enriched["qc_flags"].append("master_date_missing")
            buckets["refresh"].append(enriched)
            continue
        delta = (cand_date - master_date).days
        enriched["date_delta_days"] = delta
        if delta <= 0:
            enriched["bucket"] = "skip"
            buckets["skip"].append(enriched)
        else:
            enriched["bucket"] = "refresh"
            fields = ["C"]
            if delta >= stale_threshold_days:
                fields.extend(["J", "K", "L", "N"])
            enriched["fields_to_refresh"] = fields
            buckets["refresh"].append(enriched)
    return buckets


# ============================================================================
# Multi-Files TOC parser + chapter matcher + chapter PDF fetch
# ============================================================================

def _normalize_title(t: str) -> str:
    """Normalize a chapter title for matching (lowercase, strip whitespace/punct)."""
    if not t:
        return ""
    s = t.lower()
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\[redacted\]", "", s)
    s = re.sub(r"\(redacted\)", "", s)
    s = re.sub(r"^our\s+", "", s)
    s = re.sub(r"[^a-z0-9\u4e00-\u9fff\s\-/']+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_multi_files_toc(toc_url: str) -> List[Dict[str, str]]:
    """Fetch the Multi-Files TOC htm and return one entry per chapter PDF.

    Each entry is {'title': str (as-rendered), 'url': str (absolute)}.
    """
    r = _http_get(toc_url)
    soup = BeautifulSoup(r.text, "html.parser")
    entries: List[Dict[str, str]] = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        title = " ".join(a.get_text().split()).strip()
        if not title:
            continue
        abs_url = urljoin(toc_url, href)
        entries.append({"title": title, "url": abs_url})
    return entries


def match_target_chapters(
    toc: List[Dict[str, str]],
) -> Dict[str, Optional[Dict[str, str]]]:
    """Pick SUMMARY/BUSINESS/FINANCIAL/PARTIES/HISTORY entries via CHAPTER_VARIANTS.

    Uses a two-pass strategy: (1) exact-ish variant match, (2) difflib fuzzy
    fallback with ratio >= 0.7. PARTIES + HISTORY are best-effort: if absent
    they default to None and downstream extractors handle gracefully.
    """
    out: Dict[str, Optional[Dict[str, str]]] = {
        "summary": None, "business": None, "financial": None,
        "parties": None, "history": None,
    }
    norm_entries = [(e, _normalize_title(e["title"])) for e in toc]
    for target, variants in CHAPTER_VARIANTS.items():
        hit: Optional[Dict[str, str]] = None
        norm_variants = [_normalize_title(v) for v in variants]
        for v_norm in norm_variants:
            for entry, t_norm in norm_entries:
                if not t_norm:
                    continue
                # exact or substring match in either direction
                if t_norm == v_norm:
                    hit = entry
                    break
                if len(v_norm) >= 6 and v_norm in t_norm:
                    hit = entry
                    break
                if len(t_norm) <= 40 and t_norm in v_norm and len(t_norm) >= 6:
                    hit = entry
                    break
            if hit is not None:
                break
        if hit is None:
            best_ratio, best_entry = 0.0, None
            for entry, t_norm in norm_entries:
                if not t_norm:
                    continue
                for v_norm in norm_variants:
                    ratio = difflib.SequenceMatcher(None, t_norm, v_norm).ratio()
                    if ratio > best_ratio:
                        best_ratio, best_entry = ratio, entry
            if best_ratio >= 0.7:
                hit = best_entry
        out[target] = hit
    return out


def fetch_chapter_pdf(url: str, cache_dir: str) -> str:
    """Download a single chapter PDF into cache_dir; return local absolute path.

    Idempotent: if the target file already exists and is non-empty, returns
    the path without re-fetching. Raises HkexFetchError on non-200 or
    non-application/pdf content-type.
    """
    os.makedirs(cache_dir, exist_ok=True)
    parts = url.rstrip("/").split("/")
    fname = "_".join(parts[-2:]) if len(parts) >= 2 else parts[-1]
    local_path = os.path.join(cache_dir, fname)
    if os.path.exists(local_path) and os.path.getsize(local_path) > 1024:
        return local_path
    r = _http_get(url)
    content_type = r.headers.get("Content-Type", "").lower()
    if "pdf" not in content_type and not url.lower().endswith(".pdf"):
        raise HkexFetchError(
            "unexpected content-type {!r} at {}".format(content_type, url))
    with open(local_path, "wb") as f:
        f.write(r.content)
    return local_path


def fetch_targeted_chapters(
    candidate: Dict[str, Any],
    cache_dir: str,
) -> Dict[str, Any]:
    """Download SUMMARY/BUSINESS/FINANCIAL chapters for one candidate.

    Falls back to Full Version PDF when Multi-Files TOC is absent.
    Mutates candidate in place and also returns it.
    """
    candidate.setdefault("qc_flags", [])
    toc_url = candidate.get("multi_files_url")
    # CRITICAL slots = SUMMARY + BUSINESS + FINANCIAL (drive pdf_status).
    # OPTIONAL slots = PARTIES + HISTORY (sponsor + structure verification);
    # absence is logged but does not change pdf_status.
    CRITICAL_SLOTS = ("summary", "business", "financial")
    OPTIONAL_SLOTS = ("parties", "history")
    ALL_SLOTS = CRITICAL_SLOTS + OPTIONAL_SLOTS
    chapter_paths: Dict[str, Optional[str]] = {s: None for s in ALL_SLOTS}
    chapter_urls: Dict[str, Optional[str]] = {s: None for s in ALL_SLOTS}
    chapter_sizes: Dict[str, int] = {}
    missing: List[str] = []
    pdf_status = "ok"
    if not toc_url:
        # Fall back to Full Version
        full_url = candidate.get("full_pdf_url")
        if not full_url:
            candidate["pdf_status"] = "failed"
            candidate["qc_flags"].append("no_toc_and_no_full_pdf")
            candidate["chapter_paths"] = chapter_paths
            candidate["chapter_urls"] = chapter_urls
            candidate["chapter_sizes"] = chapter_sizes
            candidate["missing_chapters"] = list(CRITICAL_SLOTS)
            return candidate
        try:
            full_path = fetch_chapter_pdf(full_url, cache_dir)
            # Treat the full PDF as the source for ALL slots; extractors
            # paginate across it. PARTIES + HISTORY are also routed to the
            # full PDF — sponsor regex still works on the combined text.
            for slot in ALL_SLOTS:
                chapter_paths[slot] = full_path
                chapter_urls[slot] = full_url
            chapter_sizes = {
                k: os.path.getsize(full_path) for k in chapter_paths
            }
            candidate["qc_flags"].append("no_toc_fallback_full_pdf")
            pdf_status = "no_toc"
        except HkexFetchError as e:
            candidate["qc_flags"].append("full_pdf_fetch_failed: {}".format(e))
            pdf_status = "failed"
        candidate["pdf_status"] = pdf_status
        candidate["chapter_paths"] = chapter_paths
        candidate["chapter_urls"] = chapter_urls
        candidate["chapter_sizes"] = chapter_sizes
        candidate["missing_chapters"] = [k for k, v in chapter_paths.items() if not v]
        return candidate
    # Normal path: parse TOC
    try:
        toc = parse_multi_files_toc(toc_url)
    except HkexFetchError as e:
        candidate["pdf_status"] = "failed"
        candidate["qc_flags"].append("toc_fetch_failed: {}".format(e))
        candidate["chapter_paths"] = chapter_paths
        candidate["chapter_urls"] = chapter_urls
        candidate["chapter_sizes"] = chapter_sizes
        candidate["missing_chapters"] = list(CRITICAL_SLOTS)
        return candidate
    candidate["toc"] = toc
    picks = match_target_chapters(toc)
    optional_missing: List[str] = []
    for slot in ALL_SLOTS:
        entry = picks.get(slot)
        if entry is None:
            if slot in CRITICAL_SLOTS:
                missing.append(slot)
            else:
                optional_missing.append(slot)
            continue
        chapter_urls[slot] = entry["url"]
        try:
            local = fetch_chapter_pdf(entry["url"], cache_dir)
            chapter_paths[slot] = local
            chapter_sizes[slot] = os.path.getsize(local)
        except HkexFetchError as e:
            candidate["qc_flags"].append("{}_fetch_failed: {}".format(slot, e))
            if slot in CRITICAL_SLOTS:
                missing.append(slot)
            else:
                optional_missing.append(slot)
    if optional_missing:
        candidate["qc_flags"].append(
            "optional_chapters_missing: " + ",".join(optional_missing)
        )
    if missing:
        if "summary" in missing and ("business" in missing or "financial" in missing):
            pdf_status = "failed"
        else:
            pdf_status = "partial"
    candidate["pdf_status"] = pdf_status
    candidate["chapter_paths"] = chapter_paths
    candidate["chapter_urls"] = chapter_urls
    candidate["chapter_sizes"] = chapter_sizes
    candidate["missing_chapters"] = missing
    return candidate


# ============================================================================
# pdfplumber financial table extractor
# ============================================================================

_UNIT_PATTERNS = (
    (re.compile(r"RMB\s*in\s*thousands|RMB['\u2019]?\s*000", re.IGNORECASE), 0.001),
    (re.compile(r"RMB\s*in\s*millions|RMB\s*million", re.IGNORECASE), 1.0),
    (re.compile(r"RMB\s*in\s*billions|RMB\s*billion", re.IGNORECASE), 1000.0),
    (re.compile(r"US\$\s*in\s*thousands|USD\s*in\s*thousands", re.IGNORECASE), 0.001),
    (re.compile(r"US\$\s*in\s*millions|USD\s*in\s*millions", re.IGNORECASE), 1.0),
)

_FY_HEADER_RE = re.compile(
    r"(?:for\s+the\s+years?\s+ended\s+december\s*31|"
    r"as\s+of\s+december\s*31|"
    r"year\s+ended\s+december\s*31)",
    re.IGNORECASE,
)
_YEAR_TUPLE_RE = re.compile(r"(20\d\d)\s+(20\d\d)")
_NUM_RE = re.compile(r"\(?([\d,]+(?:\.\d+)?)\)?")


def _clean_page_text(text: str) -> str:
    """Strip (cid:XX) artifacts, dot leaders (consecutive or space-separated),
    and collapse whitespace. Dot leaders are common in HKEX financial tables
    (e.g., "Revenue . . . . . . 984,848 685,842") and must be normalized or
    they push the label portion past the max_label_chars guard.
    """
    if not text:
        return ""
    s = re.sub(r"\(cid:\d+\)", " ", text)
    # Collapse consecutive dots
    s = re.sub(r"\.{3,}", " ", s)
    # Collapse space-separated dot leaders (e.g. ". . . . . .")
    s = re.sub(r"(?:\s*\.\s*){3,}", " ", s)
    s = re.sub(r"\u2024+", " ", s)
    s = re.sub(r"\u2026+", " ", s)
    s = re.sub(r"[ \t]+", " ", s)
    return s


def _parse_number(s: str) -> Optional[float]:
    """Parse a number string like '(341,957)' or '1,172.5' into float."""
    s = s.strip()
    if not s or s in ("-", "\u2013", "\u2014", "nil", "Nil", "NIL"):
        return 0.0
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    s = s.replace(",", "").replace(" ", "")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if negative else v


def _find_unit_multiplier(text: str) -> Optional[float]:
    """Return multiplier to convert raw numbers to RMB millions (or USD millions)."""
    for pattern, mult in _UNIT_PATTERNS:
        if pattern.search(text):
            return mult
    return None


def _numbers_on_line(line: str) -> List[float]:
    """Extract all parseable numeric tokens from a line, excluding years and footnote refs."""
    tokens = re.findall(r"\(?-?[\d,]+(?:\.\d+)?\)?", line)
    out: List[float] = []
    for t in tokens:
        if not any(ch.isdigit() for ch in t):
            continue
        v = _parse_number(t)
        if v is None:
            continue
        out.append(v)
    # Strip footnote refs (small integers 1..9 in parentheses) first
    filtered = [v for v in out if not (abs(v) < 10 and v == int(v))]
    # Strip year-like tokens (2000..2099) if we still have >= 2 non-year numbers
    non_years = [v for v in filtered if not (2000.0 <= v <= 2099.0 and v == int(v))]
    if len(non_years) >= 2:
        return non_years
    return filtered if len(filtered) >= 2 else out


_NARRATIVE_MARKERS = (
    "increased from", "decreased from", "amounted to", "increased by",
    "decreased by", "which amounted", "for the year ended",
    "recognized loss", "recorded loss", "recorded profit",
    "was primarily due to", "mainly due to",
    "remained stable", "rmb", "million in 20", "billion in 20",
)


def _is_narrative_line(line: str) -> bool:
    """Reject lines that look like full sentences rather than tabular rows."""
    lt = line.lower()
    if any(m in lt for m in _NARRATIVE_MARKERS):
        return True
    # Too long with numbers near the end -> likely a narrative run-on
    if len(line) > 160:
        return True
    return False


def _first_number_position(line: str) -> int:
    """Return the char index of the first number-like token in a line, or -1."""
    m = re.search(r"\(?-?\d[\d,]*(?:\.\d+)?\)?", line)
    return m.start() if m else -1


def _extract_two_year_row_window(
    lines: List[str],
    value_idx: int,
    metric_pattern: re.Pattern,
    label_lookback: int = 2,
    max_label_chars: int = 80,
) -> Optional[Tuple[float, float]]:
    """Match metric label in current line or a preceding continuation line and
    extract the last two numeric tokens from the value row.

    Guards against narrative run-on sentences:
      - The value line must not look like a full sentence (narrative markers).
      - The label portion (text before the first number on the value line)
        must be no longer than max_label_chars.
      - For multi-line labels, the preceding lookback lines must have NO
        numbers of their own AND be ≤ max_label_chars in length.
    """
    value_line = lines[value_idx]
    nums = _numbers_on_line(value_line)
    if len(nums) < 2:
        return None
    if _is_narrative_line(value_line):
        return None
    first_num_pos = _first_number_position(value_line)
    if first_num_pos == -1:
        return None
    label_portion = value_line[:first_num_pos]
    if metric_pattern.search(value_line) and len(label_portion) <= max_label_chars:
        return (nums[-2], nums[-1])
    start = max(0, value_idx - label_lookback)
    for j in range(value_idx - 1, start - 1, -1):
        prev = lines[j]
        if not prev:
            continue
        if _numbers_on_line(prev):
            return None
        if len(prev) > max_label_chars:
            return None
        if _is_narrative_line(prev):
            return None
        if metric_pattern.search(prev):
            if len(label_portion) > max_label_chars:
                return None
            return (nums[-2], nums[-1])
    return None


_STRICT_LOSS_LABEL = re.compile(
    r"^\s*(?:loss|profit)\s+(?:for\s+the\s+(?:year|period)"
    r"|attributable\s+to(?:\s+(?:owners|equity\s+holders))?)",
    re.IGNORECASE,
)
_TOTAL_REVENUE_LABEL = re.compile(
    r"^\s*(?:total\s+)?revenue(?:s)?\s*$|^\s*(?:total\s+)?revenue(?:s)?\s+[\d(]",
    re.IGNORECASE,
)
_CASH_EQUIV_LABEL = re.compile(
    r"^\s*cash\s+and\s+cash\s+equivalents\s*$|"
    r"^\s*cash\s+and\s+cash\s+equivalents\s+[\d(]",
    re.IGNORECASE,
)


def _process_table_rows(
    tables: List[List[List[Optional[str]]]],
    page_num: int,
    unit_mult: float,
    result: Dict[str, Any],
) -> None:
    """Walk pdfplumber table rows and populate result with metric values."""
    for table in tables:
        for row in table:
            if not row or len(row) < 2:
                continue
            # Cells may contain None; normalize.
            cells = [
                (c or "").strip().replace("\n", " ") for c in row
            ]
            label = cells[0]
            if not label:
                continue
            # Collect numeric values from remaining cells
            values: List[float] = []
            for c in cells[1:]:
                if not c:
                    continue
                nums = _numbers_on_line(c)
                if nums:
                    values.append(nums[-1])
            if len(values) < 1:
                continue
            label_clean = label.strip()

            if result["net_income_m"] is None and _STRICT_LOSS_LABEL.search(label_clean):
                ni = values[-1]
                if re.search(r"\bloss\b", label_clean, re.IGNORECASE) and ni > 0:
                    ni = -ni
                result["net_income_m"] = round(ni * unit_mult, 3)
                result["source_pages"]["net_income"] = page_num
                result["confidence"]["net_income"] = "high"
                result["raw_hits"].append({
                    "metric": "net_income", "page": page_num,
                    "line": "{} -> {}".format(label_clean[:80], values),
                })

            if result["revenue_m"] is None and _TOTAL_REVENUE_LABEL.search(label_clean):
                result["revenue_m"] = round(values[-1] * unit_mult, 3)
                result["source_pages"]["revenue"] = page_num
                result["confidence"]["revenue"] = "high"
                result["raw_hits"].append({
                    "metric": "revenue", "page": page_num,
                    "line": "{} -> {}".format(label_clean[:80], values),
                })

            if result["cash_m"] is None and _CASH_EQUIV_LABEL.search(label_clean):
                result["cash_m"] = round(values[-1] * unit_mult, 3)
                result["source_pages"]["cash"] = page_num
                result["confidence"]["cash"] = "high"
                result["raw_hits"].append({
                    "metric": "cash", "page": page_num,
                    "line": "{} -> {}".format(label_clean[:80], values),
                })


def extract_financial_tables_pdfplumber(pdf_path: str) -> Dict[str, Any]:
    """Scan a FINANCIAL INFORMATION chapter PDF for Revenue/NI/Cash values.

    Primary strategy: use pdfplumber's extract_tables() for row-aligned data,
    since HKEX prospectus tables often merge visually when extract_text() is
    used. Falls back to the sliding-window text parser when tables are absent.

    Returns:
      revenue_m        : float or None (None means pre-revenue / not found)
      net_income_m     : float or None
      cash_m           : float or None
      fy_label         : str (e.g. 'FY25')
      currency         : 'RMB' | 'USD' | 'unknown'
      source_pages     : dict mapping metric -> page number
      confidence       : dict mapping metric -> 'high' / 'medium' / 'low'
    """
    import pdfplumber
    result: Dict[str, Any] = {
        "revenue_m": None,
        "net_income_m": None,
        "cash_m": None,
        "fy_label": None,
        "currency": "unknown",
        "source_pages": {},
        "confidence": {},
        "raw_hits": [],
    }

    revenue_labels = re.compile(
        r"^\s*(?:total\s+)?revenue(?:s)?\b", re.IGNORECASE
    )
    # Allow "Profit/(loss) for the year" AND OCR-merged "forthe year",
    # and also "loss attributable to owners of the parent / equity holders".
    loss_labels = re.compile(
        r"(?:loss|profit)[^\w\n]*for\s*the?\s*(?:year|period)|"
        r"loss\s+attributable\s+to",
        re.IGNORECASE,
    )
    cash_labels = re.compile(
        r"^\s*cash\s+and\s+cash\s+equivalents\b", re.IGNORECASE
    )

    with pdfplumber.open(pdf_path) as pdf:
        # First pass: determine target FY (most recent year in "For the Year Ended" headers)
        target_fy: Optional[int] = None
        for i, page in enumerate(pdf.pages):
            raw = page.extract_text() or ""
            text = _clean_page_text(raw)
            if _FY_HEADER_RE.search(text):
                for m in _YEAR_TUPLE_RE.finditer(text):
                    y1, y2 = int(m.group(1)), int(m.group(2))
                    target_fy = max(y1, y2) if target_fy is None else max(target_fy, max(y1, y2))
        if target_fy:
            result["fy_label"] = "FY{:02d}".format(target_fy % 100)

        # Determine currency / unit from the most common unit marker
        unit_mult = None
        currency = "unknown"
        for page in pdf.pages:
            raw = page.extract_text() or ""
            if "RMB" in raw:
                currency = "RMB"
            elif "US$" in raw or "USD" in raw:
                if currency == "unknown":
                    currency = "USD"
            if unit_mult is None:
                m = _find_unit_multiplier(raw)
                if m is not None:
                    unit_mult = m
        if unit_mult is None:
            unit_mult = 0.001  # default assumption for HK prospectuses (thousands)
            result["confidence"]["_unit"] = "low"
        result["currency"] = currency

        # Primary strategy: extract_tables() for row-aligned HKEX financial tables.
        for i, page in enumerate(pdf.pages):
            if (result["revenue_m"] is not None
                and result["net_income_m"] is not None
                and result["cash_m"] is not None):
                break
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            if tables:
                _process_table_rows(tables, i + 1, unit_mult, result)

        # Fallback strategy: sliding-window text parsing for missing metrics.
        for i, page in enumerate(pdf.pages):
            raw = page.extract_text() or ""
            text = _clean_page_text(raw)
            lines = [ln.strip() for ln in text.split("\n")]
            for v_idx, line in enumerate(lines):
                if len(line) < 6:
                    continue
                # A candidate "value row" has at least 2 numeric tokens after
                # year/footnote stripping.
                nums = _numbers_on_line(line)
                if len(nums) < 2:
                    continue

                # Net income / loss
                if result["net_income_m"] is None:
                    pair = _extract_two_year_row_window(lines, v_idx, loss_labels)
                    if pair is not None:
                        ni_raw = pair[1]
                        ctx_start = max(0, v_idx - 2)
                        ctx = " ".join(lines[ctx_start : v_idx + 1])
                        if re.search(r"\bloss\b", ctx, re.IGNORECASE) and ni_raw > 0:
                            ni_raw = -ni_raw
                        result["net_income_m"] = round(ni_raw * unit_mult, 3)
                        result["source_pages"]["net_income"] = i + 1
                        result["confidence"]["net_income"] = "high"
                        result["raw_hits"].append({
                            "metric": "net_income", "page": i + 1,
                            "line": ctx[-150:],
                        })

                # Cash and cash equivalents
                if result["cash_m"] is None:
                    pair = _extract_two_year_row_window(lines, v_idx, cash_labels)
                    if pair is not None:
                        result["cash_m"] = round(pair[1] * unit_mult, 3)
                        result["source_pages"]["cash"] = i + 1
                        result["confidence"]["cash"] = "high"
                        result["raw_hits"].append({
                            "metric": "cash", "page": i + 1, "line": line[:150],
                        })

                # Revenue
                if result["revenue_m"] is None:
                    pair = _extract_two_year_row_window(lines, v_idx, revenue_labels)
                    if pair is not None:
                        result["revenue_m"] = round(pair[1] * unit_mult, 3)
                        result["source_pages"]["revenue"] = i + 1
                        result["confidence"]["revenue"] = "high"
                        result["raw_hits"].append({
                            "metric": "revenue", "page": i + 1, "line": line[:150],
                        })

    # Pre-revenue fallback: if revenue still None and any narrative mentions
    # "have not generated any revenue" / "pre-revenue", mark as dash
    if result["revenue_m"] is None:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                raw = (page.extract_text() or "").lower()
                if ("have not generated" in raw and "revenue" in raw) or \
                   "no revenue" in raw or \
                   "pre-revenue" in raw:
                    result["revenue_m"] = "-"
                    result["source_pages"]["revenue"] = 0
                    result["confidence"]["revenue"] = "medium"
                    break

    # Sanity check: detect unit-mismatch (e.g., revenue=305m but NI=43199m
    # would be inconsistent — unit_mult was likely applied to one but not
    # the other). When all three are populated and any pair differs by >100x
    # in absolute magnitude, downgrade confidence on the outlier.
    _sanity_check_unit_consistency(result)

    return result


# ============================================================================
# Narrative prose fallback for revenue / NI / cash
# ============================================================================

# Anchor regex (locate the START of a narrative metric statement)
# then we scan forward up to 250 chars for ALL "RMB X million" tokens and take
# the LAST one -- because TRP statements are always written oldest-to-newest:
#   "In 2023, 2024 and 2025, we recorded revenue of RMB495.9 million,
#    RMB544.1 million and RMB701.5 million"  -> we want 701.5 (FY25).
_NARRATIVE_REV_ANCHOR = re.compile(
    r"(?:total\s+)?revenue\s+(?:of|was|"
    r"amounted\s+to|increased\s+to|increased\s+by\s+\S+\s+to|"
    r"of\s+approximately|grew\s+to|reached|"
    r"declined\s+to|decreased\s+to|dropped\s+to)",
    re.IGNORECASE,
)
# For NI / loss anchor: accept the narrow IFRS forms (preferred) AND broader
# "loss before tax(ation)" / "loss from operations" since clinical-stage
# biotechs often disclose losses at those P&L lines. We bias toward the LAST
# anchor so a paragraph that mentions both "loss from operations 142m" then
# "loss for the year 191m" picks the latter (which is what we actually want).
_NARRATIVE_LOSS_ANCHOR = re.compile(
    r"(?:loss|profit)\s+(?:"
    r"for\s+the\s+(?:year|period)"
    r"|attributable"
    r"|before\s+tax(?:ation)?"
    r"|from\s+operations"
    r")",
    re.IGNORECASE,
)
_NARRATIVE_CASH_ANCHOR = re.compile(
    r"cash\s+and\s+cash\s+equivalents\s+(?:at\s+(?:the\s+)?end\s+of"
    r"|of\s+approximately|amounted\s+to|was|of\s+|increased\s+to)",
    re.IGNORECASE,
)
# After the anchor: capture all RMB X million tokens in next 250 chars
_RMB_MILLION_TOKEN = re.compile(
    r"(?:RMB|HK\$|US\$|USD)\s*([\d,]+(?:\.\d+)?)\s*(million|billion)",
    re.IGNORECASE,
)


def _is_loss_in_context(text_window: str) -> bool:
    """Return True if narrative within ±100 chars of match indicates a loss."""
    return bool(re.search(r"\bloss\b", text_window, re.IGNORECASE))


def _extract_last_rmb_value(text_after_anchor: str, max_chars: int = 250) -> Optional[float]:
    """From a slice starting at a metric anchor, return the LAST RMB X million
    token converted to RMB millions. Handles 'million' (×1) / 'billion' (×1000).
    Stops scanning at the first sentence-end (period or newline followed by
    whitespace OR start of next paragraph), so we don't spill into the next
    metric's discussion.
    """
    slice_text = text_after_anchor[:max_chars]
    # Truncate at sentence end OR newline-paragraph boundary
    sent_end = re.search(
        r"\.(?:\s+[A-Za-z])"      # period + space + any letter (sentence)
        r"|(?:\.\s*\n)"             # period + newline (sentence break)
        r"|(?:\n\s*\n)"             # blank line (paragraph break)
        r"|(?:\n[A-Z])"             # newline + capital (new paragraph header)
        r"|(?:respectively\.)",     # explicit list-end marker
        slice_text,
    )
    if sent_end:
        slice_text = slice_text[: sent_end.start() + 1]
    matches = list(_RMB_MILLION_TOKEN.finditer(slice_text))
    if not matches:
        return None
    last = matches[-1]
    try:
        v = float(last.group(1).replace(",", ""))
        unit = last.group(2).lower()
        if unit.startswith("billion"):
            v *= 1000.0
        return v
    except ValueError:
        return None


def extract_financials_from_narrative(text: str) -> Dict[str, Any]:
    """Extract J/K/L from prose narrative when table extraction fails.

    Targets statements like:
      - "In 2023, 2024 and 2025, we recorded total revenue of RMB495.9 million,
         RMB544.1 million and RMB701.5 million" -> 701.5 (last/most-recent)
      - "loss for the year was RMB163.8 million and RMB191.4 million" -> -191.4
      - "cash and cash equivalents at end of year amounted to RMB268.7 million"
        -> 268.7

    Returns the same shape as extract_financial_tables_pdfplumber. All values
    in RMB millions (anchor regex requires 'million'/'billion' suffix; billion
    auto-converts to millions).
    """
    result: Dict[str, Any] = {
        "revenue_m": None, "net_income_m": None, "cash_m": None,
        "fy_label": None, "currency": "RMB",
        "source_pages": {}, "confidence": {}, "raw_hits": [],
    }
    if not text:
        return result

    cleaned = _clean_page_text(text)

    # Revenue: find LAST anchor (latest narrative paragraph) and take LAST
    # RMB-million token after it
    rev_anchors = list(_NARRATIVE_REV_ANCHOR.finditer(cleaned))
    if rev_anchors:
        anchor = rev_anchors[-1]
        v = _extract_last_rmb_value(cleaned[anchor.end():])
        if v is not None:
            result["revenue_m"] = v
            result["confidence"]["revenue"] = "medium"
            result["source_pages"]["revenue"] = -1
            result["raw_hits"].append({
                "metric": "revenue", "page": -1,
                "line": cleaned[max(0, anchor.start() - 20): anchor.end() + 200][:300],
            })

    # Net income / loss
    pl_anchors = list(_NARRATIVE_LOSS_ANCHOR.finditer(cleaned))
    if pl_anchors:
        anchor = pl_anchors[-1]
        v = _extract_last_rmb_value(cleaned[anchor.end():])
        if v is not None:
            ctx = cleaned[max(0, anchor.start() - 100): anchor.end() + 250]
            if _is_loss_in_context(ctx):
                v = -abs(v)
            result["net_income_m"] = v
            result["confidence"]["net_income"] = "medium"
            result["source_pages"]["net_income"] = -1
            result["raw_hits"].append({
                "metric": "net_income", "page": -1,
                "line": ctx[:300],
            })

    # Cash
    cash_anchors = list(_NARRATIVE_CASH_ANCHOR.finditer(cleaned))
    if cash_anchors:
        anchor = cash_anchors[-1]
        v = _extract_last_rmb_value(cleaned[anchor.end():])
        if v is not None:
            result["cash_m"] = v
            result["confidence"]["cash"] = "medium"
            result["source_pages"]["cash"] = -1
            result["raw_hits"].append({
                "metric": "cash", "page": -1,
                "line": cleaned[max(0, anchor.start() - 20): anchor.end() + 100][:200],
            })

    return result


def _sanity_check_unit_consistency(result: Dict[str, Any]) -> None:
    """Flag unit-mismatch when J/K/L magnitudes are inconsistent.

    HKEX prospectuses always report revenue/NI/cash in the same unit (RMB'000
    or RMB millions). If after unit-mult application revenue=305m but NI=43199
    (i.e., NI is ~140x larger), the unit_mult was applied to revenue but not
    NI — most likely the NI value came from a different page with different
    unit metadata. Downgrade confidence on the outlier so Phase 4 surfaces it.
    """
    j = result.get("revenue_m")
    k = result.get("net_income_m")
    l = result.get("cash_m")
    nums = [(name, abs(v)) for name, v in
            (("revenue", j), ("net_income", k), ("cash", l))
            if isinstance(v, (int, float)) and v != 0]
    if len(nums) < 2:
        return
    # If any pair differs by >500x, flag the largest as suspect
    nums.sort(key=lambda x: x[1])
    smallest = nums[0][1]
    for name, v in nums[1:]:
        if smallest > 0 and v / smallest > 500:
            existing = result["confidence"].get(name, "high")
            if existing != "low":
                result["confidence"][name] = "low"
                result.setdefault("sanity_flags", []).append(
                    "{}_unit_mismatch:{:.1f}x_vs_smallest".format(name, v / smallest)
                )


# ============================================================================
# Per-field extractors and top-level row_draft builder
# ============================================================================

def _extract_chapter_text(pdf_path: str, max_pages: int = 10) -> str:
    """Return cleaned text of the first max_pages of a chapter PDF."""
    import pdfplumber
    parts: List[str] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= max_pages:
                    break
                parts.append(_clean_page_text(page.extract_text() or ""))
    except Exception as e:
        warnings_log.append("chapter_text_read_failed: {} ({})".format(pdf_path, e))
    return "\n".join(parts)


# Known sponsor-bank names/aliases. Used to map a free-text sponsor block to
# a canonical short label set (e.g., "China International Capital Corporation
# Hong Kong Securities Limited" → "CICC"). Order matters for substring matching.
_SPONSOR_BANK_ALIASES: List[Tuple[str, str]] = [
    ("china international capital corporation", "CICC"),
    ("china merchants securities (hk)", "CMS HK"),
    ("china merchants securities", "CMS HK"),
    ("merrill lynch", "BAML"),
    ("bofa securities", "BAML"),
    ("bank of america", "BAML"),
    ("ccb international", "CCBI"),
    ("citic securities", "CITIC"),
    ("guotai junan", "Guotai Junan"),
    ("guotai capital", "Guotai Junan"),
    ("haitong international", "Haitong"),
    ("huatai financial", "Huatai"),
    ("huatai securities", "Huatai"),
    ("jefferies", "Jefferies"),
    ("morgan stanley", "Morgan Stanley"),
    ("goldman sachs", "Goldman Sachs"),
    ("ubs", "UBS"),
    ("credit suisse", "CS"),
    ("nomura", "Nomura"),
    ("daiwa", "Daiwa"),
    ("gf capital", "GF Capital"),
    ("gf securities", "GF Capital"),
    ("abci capital", "ABCI"),
    ("icbc international", "ICBCI"),
    ("boci asia", "BOCI"),
    ("boc international", "BOCI"),
    ("clsa limited", "CLSA"),
    ("hsbc", "HSBC"),
    ("standard chartered", "StanChart"),
    ("deutsche bank", "DB"),
]


def _normalize_sponsor_block(text: str) -> str:
    """Map a sponsor block text to a comma-separated canonical-alias string.

    Falls back to the trimmed raw text when no known aliases match.
    """
    if not text:
        return ""
    found: List[str] = []
    text_lower = text.lower()
    for needle, alias in _SPONSOR_BANK_ALIASES:
        if needle in text_lower and alias not in found:
            found.append(alias)
    if found:
        return ", ".join(found)
    # Fallback: trim raw text to first comma/parenthesis
    cleaned = re.sub(r"\s+", " ", text).strip(" .;:,")
    for end in [". The", "\n", "and (the", "(Act", "(the", "in its", "in their"]:
        if end in cleaned:
            cleaned = cleaned.split(end)[0]
    return cleaned[:200]


def _extract_col_I_sponsor(parties_text: str) -> Optional[str]:
    """Pull sponsor bank names from the PARTIES INVOLVED section.

    Note: pre-May-2026 versions of this function read the SUMMARY chapter, but
    sponsor info is consistently in 'DIRECTORS AND PARTIES INVOLVED' (or its
    A-share variant 'DIRECTORS, SUPERVISORS AND PARTIES INVOLVED'). When no
    PARTIES chapter is available (rare — TOC missing or full PDF fallback),
    the caller should pass the full SUMMARY text and the patterns below will
    still fire on any 'Sole/Joint Sponsor:' label they happen to find.
    """
    if not parties_text:
        return None
    # Look for sponsor section header followed by 1500 chars of bank listings.
    # The PARTIES chapter has structure: "PARTIES INVOLVED IN THE [REDACTED]\n
    # Sole Sponsor   <bank A>\n   <address>\nLegal advisers..." — we want the
    # block between "Sole Sponsor"/"Joint Sponsor(s)" and "Legal advis".
    section_re = re.compile(
        r"(?:Sole\s+Sponsor|Joint\s+Sponsors?(?:\s*,\s*Sponsor-OCs?)?|Sponsor-OCs?)"
        r"[\s\.,:]{0,40}([\s\S]{20,2000}?)"
        r"(?=Legal\s+Advis|Reporting\s+Accountant|Compliance\s+Adviser|–\s*\d+\s*–|$)",
        re.IGNORECASE,
    )
    m = section_re.search(parties_text)
    if m:
        block = m.group(1)
        normalized = _normalize_sponsor_block(block)
        if normalized:
            return normalized
    # Single-line fallback (older/simpler templates)
    patterns = [
        r"Sole\s+Sponsor[\s:\.]*(?:\([^)]+\))?[\s:\.]*([^\n]{5,300})",
        r"Joint\s+Sponsors?[\s:\.]*(?:\([^)]+\))?[\s:\.]*([^\n]{5,300})",
        r"Sponsor\(s\)[\s:\.]*([^\n]{5,300})",
    ]
    for p in patterns:
        m = re.search(p, parties_text, re.IGNORECASE)
        if m:
            normalized = _normalize_sponsor_block(m.group(1))
            if normalized and 3 <= len(normalized) <= 200:
                return normalized
    return None


def extract_fields_from_chapters(
    candidate: Dict[str, Any],
    target_fy: str = "FY25",
) -> Dict[str, Any]:
    """Build a row_draft dict for one candidate from its cached chapter PDFs.

    Responsibility split:
      - Col C/D/E          from HKEX feed (high confidence)
      - Col I (sponsor)    from pdfplumber + regex on SUMMARY (deterministic)
      - Col J/K/L          from pdfplumber table parser on FINANCIAL chapter
      - Col N              assembled from FY label + pdf_status flags
      - Col F/G/H/M        LEFT AS None HERE -- filled later by the skill
                           driver via mcp__firecrawl__scrape and merged in
                           through `apply_firecrawl_narrative()`. Each of
                           these four cols is tagged with a
                           `firecrawl_pending_col_X` QC flag so that a run
                           with Firecrawl skipped surfaces them loudly in
                           the Phase 4 diff.
    """
    row_draft: Dict[str, Any] = {
        "C": candidate.get("filing_date"),
        "D": candidate.get("name_en"),
        "E": candidate.get("name_cn"),
    }
    provenance: Dict[str, str] = {
        "C": "hkex_feed",
        "D": "hkex_feed",
        "E": "hkex_cn_feed" if candidate.get("name_cn") else "missing",
    }
    confidence: Dict[str, str] = {
        "C": "high",
        "D": "high",
        "E": "high" if candidate.get("name_cn") else "low",
    }
    qc_flags: List[str] = list(candidate.get("qc_flags", []))

    paths = candidate.get("chapter_paths") or {}
    summary_path = paths.get("summary")
    financial_path = paths.get("financial")
    parties_path = paths.get("parties")

    # Regex-based sponsor extraction now reads the PARTIES chapter (where
    # sponsor info actually lives). Falls back to SUMMARY only if PARTIES is
    # missing — older code path which we keep as a safety net but expect to
    # rarely fire.
    parties_text = (
        _extract_chapter_text(parties_path, max_pages=8) if parties_path else ""
    )
    summary_text = (
        _extract_chapter_text(summary_path, max_pages=10) if summary_path else ""
    )

    # F/G/H/M: deferred to Firecrawl (apply_firecrawl_narrative).
    for col in ("F", "G", "H", "M"):
        row_draft[col] = None
        provenance[col] = "firecrawl_pending"
        confidence[col] = "low"
        qc_flags.append("firecrawl_pending_col_{}".format(col))

    # I: sponsor via regex on PARTIES (preferred) or SUMMARY fallback.
    sponsor_source = "PARTIES" if parties_text else "SUMMARY"
    sponsor_text = parties_text or summary_text
    row_draft["I"] = _extract_col_I_sponsor(sponsor_text)
    provenance["I"] = "pdfplumber:{}".format(sponsor_source)
    confidence["I"] = "medium" if row_draft["I"] else "low"
    if not row_draft["I"]:
        qc_flags.append("sponsor_extraction_failed")

    # J/K/L: 3-tier extraction strategy.
    #   Tier 1 -- pdfplumber.extract_tables() on FINANCIAL chapter (clean tables)
    #   Tier 2 -- text-based sliding window on FINANCIAL (fallback for visual tables)
    #   Tier 3 -- narrative prose extraction on SUMMARY chapter (rescue when
    #             tier 1 + 2 return wrong values due to dotted-leader / cid:2
    #             artifacts; SUMMARY's "Our Financial Performance" narrative
    #             section explicitly says "we recorded total revenue of RMB X
    #             million" which is unambiguous)
    fin_fy_label = ""
    if financial_path:
        fin = extract_financial_tables_pdfplumber(financial_path)
        rev = fin.get("revenue_m")
        row_draft["J"] = rev if rev is not None else "-"
        row_draft["K"] = fin.get("net_income_m")
        row_draft["L"] = fin.get("cash_m")
        provenance["J"] = "pdfplumber:FINANCIAL"
        provenance["K"] = "pdfplumber:FINANCIAL"
        provenance["L"] = "pdfplumber:FINANCIAL"
        fin_conf = fin.get("confidence", {})
        confidence["J"] = "medium" if row_draft["J"] == "-" else fin_conf.get("revenue", "low")
        confidence["K"] = fin_conf.get("net_income", "low")
        confidence["L"] = fin_conf.get("cash", "low")
        fin_fy_label = fin.get("fy_label") or ""
        if fin.get("sanity_flags"):
            qc_flags.extend(fin["sanity_flags"])

        # Tier 3: narrative-prose rescue. Triggers when any J/K/L is None OR
        # has confidence='low'. Reads SUMMARY chapter "Our Financial
        # Performance" / "OVERVIEW" section.
        needs_rescue = (
            row_draft["J"] in (None, "-")
            or row_draft["K"] is None
            or row_draft["L"] is None
            or confidence["J"] == "low"
            or confidence["K"] == "low"
            or confidence["L"] == "low"
        )
        if needs_rescue and summary_text:
            narr = extract_financials_from_narrative(summary_text)
            for col, key in (("J", "revenue_m"), ("K", "net_income_m"), ("L", "cash_m")):
                narr_val = narr.get(key)
                cur_val = row_draft[col]
                cur_conf = confidence[col]
                # Only overwrite when narrative confidence is medium AND
                # current is None or low-confidence
                if (narr_val is not None
                        and (cur_val is None
                             or (col == "J" and cur_val == "-")
                             or cur_conf == "low")):
                    row_draft[col] = narr_val
                    provenance[col] = "narrative:SUMMARY"
                    confidence[col] = narr.get("confidence", {}).get(key, "medium")
    else:
        row_draft["J"] = row_draft["K"] = row_draft["L"] = None
        provenance["J"] = provenance["K"] = provenance["L"] = "missing"
        confidence["J"] = confidence["K"] = confidence["L"] = "low"
        qc_flags.append("financial_chapter_missing")

    highlights: List[str] = []
    if fin_fy_label and target_fy and fin_fy_label < target_fy:
        highlights.append("{} data; {} pending.".format(fin_fy_label, target_fy))
    if candidate.get("pdf_status") and candidate["pdf_status"] != "ok":
        highlights.append("[PHASE0 pdf_status={}]".format(candidate["pdf_status"]))
    row_draft["N"] = " ".join(highlights) if highlights else None
    provenance["N"] = "assembled"
    confidence["N"] = "high"

    return {
        "target_bucket": candidate.get("bucket", "new"),
        "master_row_idx": candidate.get("master_row_idx"),
        "fields_to_refresh": candidate.get("fields_to_refresh", []),
        "row_draft": row_draft,
        "candidate": candidate,
        "pdf_status": candidate.get("pdf_status", "ok"),
        "_provenance": provenance,
        "_confidence": confidence,
        "_qc_flags": qc_flags,
    }


def apply_firecrawl_narrative(
    staging_row: Dict[str, Any],
    fc_data: Dict[str, Any],
    source_label: str = "firecrawl:SUMMARY",
) -> Dict[str, Any]:
    """Merge a Firecrawl /scrape JSON-format result into a staging_row.

    The skill driver is expected to call mcp__firecrawl__scrape with
    `FIRECRAWL_NARRATIVE_SCHEMA` / `FIRECRAWL_NARRATIVE_PROMPT` on the
    candidate's SUMMARY chapter URL (found at
    `staging_row['candidate']['chapter_urls']['summary']`), then pass the
    returned `json` dict in here as `fc_data`.

    Non-null values from Firecrawl overwrite F/G/H/M in row_draft, update
    provenance to `source_label`, mark confidence as 'high', and clear the
    matching `firecrawl_pending_col_X` QC flags. Fields Firecrawl returned
    null for stay as None with the pending flag intact -- Phase 4 will
    surface them so the user can fill manually.
    """
    row_draft = staging_row["row_draft"]
    provenance = staging_row["_provenance"]
    confidence = staging_row["_confidence"]
    qc_flags = staging_row["_qc_flags"]

    if not isinstance(fc_data, dict):
        return staging_row

    mapping = (
        ("F", "shareholder_structure"),
        ("G", "business_model"),
        ("H", "sector"),
        ("M", "lead_asset"),
    )

    rescued: List[str] = []
    for col, key in mapping:
        value = fc_data.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
        # Guard against LLM returning off-enum values for H.
        if col == "H" and value not in CANONICAL_SECTORS:
            qc_flags.append("firecrawl_off_enum_H: {}".format(value)[:120])
            continue
        row_draft[col] = value
        provenance[col] = source_label
        confidence[col] = "high"
        rescued.append(col)

    if rescued:
        staging_row["_qc_flags"] = [
            f for f in qc_flags
            if not any(
                f == "firecrawl_pending_col_{}".format(c) for c in rescued
            )
        ]
    return staging_row


def apply_fg_robustness_tag(
    staging_row: Dict[str, Any],
    tag: str,
    suffix_override: Optional[str] = None,
) -> str:
    """Apply (or replace) the F+G robustness tag on a staging row.

    Mutates `staging_row`:
    1. Removes any pre-existing robustness tag from `_qc_flags` (so a row
       can be re-classified, e.g., Phase 0 first sets verified_fg_prospectus
       then Phase 3 web fallback overrides to single_source_family_fg).
    2. Appends `tag` to `_qc_flags`.
    3. Strips any existing robustness suffix from `row_draft["N"]` and
       appends the canonical suffix for `tag` (or `suffix_override` when
       provided -- mainly for not_found_fg with a custom searched-list,
       or for conflicting_fg with a discarded-source footnote).

    Args:
        staging_row: Phase 0 / Phase 3 staging row dict (must contain
            `row_draft`, `_qc_flags`).
        tag: One of FG_ROBUSTNESS_TAGS. ValueError if not.
        suffix_override: Optional custom col N suffix. If None, uses
            FG_ROBUSTNESS_COL_N_SUFFIX[tag].

    Returns:
        The applied tag string (same as `tag`).

    Raises:
        ValueError: If `tag` is not in FG_ROBUSTNESS_TAGS.
    """
    if tag not in FG_ROBUSTNESS_TAGS:
        raise ValueError(
            "Invalid F+G robustness tag {!r}. Must be one of: {}".format(
                tag, FG_ROBUSTNESS_TAGS
            )
        )

    qc_flags = staging_row.get("_qc_flags", [])
    # Drop any pre-existing robustness tag, then add the new one.
    staging_row["_qc_flags"] = [f for f in qc_flags if f not in FG_ROBUSTNESS_TAGS]
    staging_row["_qc_flags"].append(tag)

    # Update col N: strip any existing robustness suffix, append the new one.
    row_draft = staging_row["row_draft"]
    existing_n = row_draft.get("N") or ""
    cleaned_n = _FG_ROBUSTNESS_SUFFIX_RE.sub("", existing_n).rstrip()
    new_suffix = suffix_override if suffix_override else FG_ROBUSTNESS_COL_N_SUFFIX[tag]
    if cleaned_n:
        row_draft["N"] = "{} {}".format(cleaned_n, new_suffix)
    else:
        row_draft["N"] = new_suffix

    return tag


def auto_classify_fg_robustness(staging_row: Dict[str, Any]) -> str:
    """Auto-classify F+G robustness for a Phase 0 prospectus-extraction row.

    Reads `_provenance` / `_confidence` / `row_draft` for cols F and G and
    picks one of three Phase-0-determinable tags:

    - `verified_fg_prospectus`: both F and G filled from "firecrawl:..."
      provenance with both `_confidence == "high"`.
    - `single_source_prospectus_fg`: both F and G filled from prospectus
      but at least one has `_confidence != "high"` (low/medium/missing).
    - `not_found_fg`: F or G is None (Firecrawl returned null and Phase 0
      finished without a fallback). The skill driver may later re-classify
      this row to web_cross_checked_fg / single_source_family_fg /
      conflicting_fg via apply_fg_robustness_tag() once Phase 3 web
      fallback completes.

    Web-fallback tags (web_cross_checked_fg / single_source_family_fg /
    conflicting_fg) are NEVER returned by this auto-classifier -- they
    require agent-side knowledge of web-fallback outcomes that the Python
    module does not see.

    Mutates `staging_row` by applying the chosen tag (delegates to
    apply_fg_robustness_tag).

    Returns:
        The applied tag string.
    """
    row_draft = staging_row["row_draft"]
    provenance = staging_row.get("_provenance", {}) or {}
    confidence = staging_row.get("_confidence", {}) or {}

    f_value = row_draft.get("F")
    g_value = row_draft.get("G")

    # Phase 0 result: F or G is None -> Firecrawl returned null and the
    # skill driver hasn't run web fallback yet (or chose not to). Tag as
    # not_found_fg; agent can override after Phase 3 if it gathers new data.
    if f_value is None or g_value is None:
        return apply_fg_robustness_tag(staging_row, "not_found_fg")

    f_prov = provenance.get("F") or ""
    g_prov = provenance.get("G") or ""
    f_from_prospectus = f_prov.startswith("firecrawl:")
    g_from_prospectus = g_prov.startswith("firecrawl:")
    f_high = confidence.get("F") == "high"
    g_high = confidence.get("G") == "high"

    if f_from_prospectus and g_from_prospectus and f_high and g_high:
        return apply_fg_robustness_tag(staging_row, "verified_fg_prospectus")

    if f_from_prospectus and g_from_prospectus:
        # Both came from prospectus but at least one isn't high-confidence.
        return apply_fg_robustness_tag(staging_row, "single_source_prospectus_fg")

    # F or G came from somewhere other than prospectus extraction. The
    # auto-classifier can't determine whether that source was cross-checked
    # or single-family -- conservative default is single_source_prospectus_fg
    # (single source, treat as needs-verify). Skill driver should override
    # via apply_fg_robustness_tag() once it knows the actual fallback path.
    return apply_fg_robustness_tag(staging_row, "single_source_prospectus_fg")


# ============================================================================
# Transient checkpoint cache management
# ============================================================================

# Transient PDF cache directory. Override via env var A1_CHECKPOINT_DIR;
# default is `hkex-a1-pipeline/checkpoints/` (sibling of the scripts/ dir).
# The repo's .gitignore excludes `checkpoints/` so downloaded PDFs never
# get committed.
CHECKPOINTS_ROOT = os.environ.get(
    "A1_CHECKPOINT_DIR",
    os.path.abspath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "checkpoints",
    )),
)


def create_cache_dir(
    run_date: Optional[datetime.date] = None,
    root: str = CHECKPOINTS_ROOT,
) -> str:
    """Create the per-run cache directory under checkpoints/ and return its path.

    The default layout is `{root}/{YYYY-MM-DD}_a1_hkex_pdfs/`. On collision
    (e.g., resuming after an aborted earlier run), a `_2`, `_3`, ... suffix is
    appended so we never clobber an existing cache.
    """
    if run_date is None:
        run_date = datetime.date.today()
    base = os.path.join(
        root,
        "{}_a1_hkex_pdfs".format(run_date.isoformat()),
    )
    path = base
    suffix = 2
    while os.path.exists(path):
        path = "{}_{}".format(base, suffix)
        suffix += 1
        if suffix > 99:
            raise RuntimeError("too many cache collisions under {}".format(root))
    os.makedirs(path, exist_ok=True)
    manifest_path = os.path.join(path, "_manifest.json")
    try:
        import json
        with open(manifest_path, "w") as f:
            json.dump(
                {
                    "created": datetime.datetime.now().isoformat(),
                    "root": root,
                    "candidates": [],
                },
                f,
                indent=2,
            )
    except Exception as e:
        warnings_log.append("manifest_write_failed: {}".format(e))
    return path


def cleanup_cache_dir(
    cache_dir: str,
    had_failures: bool = False,
) -> Optional[str]:
    """Remove the cache dir on success; preserve it on failure and return the path.

    Safety: refuses to remove anything that isn't a direct child of
    CHECKPOINTS_ROOT (or whatever was used in create_cache_dir). This protects
    against accidental rmtree of unrelated paths if a caller passes junk.
    """
    if not cache_dir or not os.path.exists(cache_dir):
        return None
    abs_target = os.path.abspath(cache_dir)
    abs_root = os.path.abspath(CHECKPOINTS_ROOT)
    if not abs_target.startswith(abs_root + os.sep):
        warnings_log.append(
            "cleanup_cache_dir refused: {} is not under {}".format(
                abs_target, abs_root))
        return abs_target
    if had_failures:
        print("Phase 0 had failures; preserving cache for debug: {}".format(abs_target))
        return abs_target
    try:
        shutil.rmtree(abs_target)
    except Exception as e:
        warnings_log.append("rmtree_failed: {} ({})".format(abs_target, e))
        return abs_target
    return None


# ============================================================================
# Standalone CLI for step-by-step testing
# ============================================================================

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print(
            "Usage: python hkex_scraper.py <year> [master_tracker.xlsx]\n"
            "  <year>          HKEX calendar year to scan, e.g. 2026\n"
            "  master_tracker  Optional path to an existing tracker xlsx;\n"
            "                  if omitted, runs feed + filter + classify\n"
            "                  against an empty master (all candidates are NEW)."
        )
        sys.exit(1)
    year = int(sys.argv[1])
    master_path = sys.argv[2] if len(sys.argv) > 2 else None
    print("Fetching HKEX feed for year={}".format(year))
    cands = fetch_lifesci_candidates(year)
    print("Total AP records: {}".format(len(cands)))
    passed, dropped = filter_candidates(cands)
    rate = len(passed) / max(len(cands), 1)
    print("Passed pre-filter: {} ({:.1%})".format(len(passed), rate))
    print("Dropped:           {}".format(len(dropped)))
    print()
    if master_path:
        print("Loading master tracker: {}".format(os.path.basename(master_path)))
        master = load_master_tracker(master_path)
        print("Master rows: {}".format(len(master["rows"])))
        print("Master A1 metadata: {}".format(master["a1_metadata"]))
    else:
        print("No master tracker provided; treating all candidates as NEW.")
        master = {"rows": [], "a1_metadata": {}, "path": None,
                  "sheet_title": None, "header_row": 2, "data_start_row": 3}
    print()
    print("Classifying candidates against master...")
    buckets = classify_candidates(passed, master)
    print("NEW     : {}".format(len(buckets["new"])))
    print("REFRESH : {}".format(len(buckets["refresh"])))
    print("SKIP    : {}".format(len(buckets["skip"])))
    print()
    for bname in ("new", "refresh", "skip"):
        items = buckets[bname]
        if not items:
            continue
        print("--- {} ({}) ---".format(bname.upper(), len(items)))
        for c in items[:10]:
            line = "  {}  id={}  {}".format(
                c["filing_date"], c["id"], c["name_en"][:60]
            )
            if bname == "refresh":
                line += "  [master r{}, delta {}d, refresh={}]".format(
                    c.get("master_row_idx"),
                    c.get("date_delta_days"),
                    "+".join(c.get("fields_to_refresh", [])),
                )
            elif bname == "skip":
                line += "  [master r{}]".format(c.get("master_row_idx"))
            print(line)
            if c.get("qc_flags"):
                print("    flags: {}".format(", ".join(c["qc_flags"])))
        if len(items) > 10:
            print("  ... ({} more)".format(len(items) - 10))
        print()
    if warnings_log:
        print("Warnings:")
        for w in warnings_log:
            print("  - {}".format(w))


# ============================================================================
# Master diff (#3) -- detect stale master cells against fresh extraction
# ============================================================================

# Cols that can drift (sponsor change, asset rename, sector reclass, etc.) and
# are worth comparing master vs extracted. Skip C (always refreshed) and N
# (always rewritten) and the financial cols J/K/L (refresh contract handles).
_DIFF_COMPARABLE_COLS = ("F", "G", "H", "I", "M")


def _normalize_for_compare(s: Any) -> str:
    """Lowercase, strip whitespace, drop punctuation noise for delta detection."""
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = s.lower()
    s = re.sub(r"[^\w一-鿿\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def diff_against_master(
    staging_row: Dict[str, Any],
    master_row: Dict[str, Any],
    cols: Tuple[str, ...] = _DIFF_COMPARABLE_COLS,
) -> List[Dict[str, Any]]:
    """Compare extracted row_draft against master cells; surface deltas.

    Used in Phase 4 (Gate 2) to flag potentially-stale master values that the
    REFRESH pipeline would otherwise preserve under the "only fill blanks" rule.
    Each delta becomes a user-decision item: keep master / take extracted /
    skip review.

    Args:
        staging_row: dict with 'row_draft' key (output of
                     extract_fields_from_chapters + apply_firecrawl_narrative).
        master_row : dict keyed by col letter (output of load_master_tracker
                     with the matched row pulled out).
        cols       : which cols to compare (default: F, G, H, I, M).

    Returns:
        list of {"col": "M", "master": <old>, "extracted": <new>,
                 "decision": "pending"} dicts -- one per non-trivial delta.
        Empty list when nothing diverges.
    """
    row_draft = staging_row.get("row_draft", {})
    deltas: List[Dict[str, Any]] = []
    for col in cols:
        master_val = master_row.get(col)
        extracted_val = row_draft.get(col)
        # Skip when either side is empty (fill-if-empty already handles this)
        if not master_val or not extracted_val:
            continue
        # Skip when normalized strings match (semantically same)
        if _normalize_for_compare(master_val) == _normalize_for_compare(extracted_val):
            continue
        # Skip if extracted is a strict prefix/suffix of master (the longer
        # string is just more detailed -- not a contradiction)
        m_norm = _normalize_for_compare(master_val)
        e_norm = _normalize_for_compare(extracted_val)
        if m_norm in e_norm or e_norm in m_norm:
            # One is a substring of the other -- prefer the longer one but
            # flag for awareness (analyst may want to take the longer one)
            if len(e_norm) > len(m_norm) * 1.3:
                deltas.append({
                    "col": col,
                    "master": str(master_val)[:200],
                    "extracted": str(extracted_val)[:200],
                    "kind": "extracted_more_detailed",
                    "decision": "pending",
                })
            continue
        deltas.append({
            "col": col,
            "master": str(master_val)[:200],
            "extracted": str(extracted_val)[:200],
            "kind": "divergent",
            "decision": "pending",
        })
    return deltas


# ============================================================================
# LLM relevance check (#4) -- detect rebrand-out-of-LS candidates (Wanchen-style)
# ============================================================================

# Schema for Firecrawl-backed LS relevance verification. The skill driver
# passes this verbatim to mcp__firecrawl__firecrawl_scrape jsonOptions for the
# Multi-Files INDEX page (NOT a chapter PDF; the index is HTML and Firecrawl
# scrapes it cleanly). For each pre-filter-passing candidate the LLM returns
# whether the company's CURRENT business (not legacy name) is in life sciences.
LLM_RELEVANCE_SCHEMA = {
    "type": "object",
    "properties": {
        "is_lifesci": {
            "type": "boolean",
            "description": (
                "True if the company's CURRENT main business is in life "
                "sciences (pharma, biotech, medtech, diagnostics, vaccines, "
                "TCM, healthcare services, healthcare tech, consumer health, "
                "CDMO/CXO). False if they have rebranded out of LS (e.g., "
                "Wanchen rebranded from Wanchen Biotechnology to Wanchen Food "
                "Group -- legacy name has 'biotech' but current business is "
                "frozen food)."
            ),
        },
        "current_business": {
            "type": "string",
            "description": "One-sentence description of the issuer's CURRENT main business.",
        },
        "rebrand_concern": {
            "type": "boolean",
            "description": (
                "True if the company has rebranded / pivoted out of life "
                "sciences (legacy name suggests LS but current business does "
                "not). False otherwise."
            ),
        },
    },
    "required": ["is_lifesci", "current_business", "rebrand_concern"],
}

LLM_RELEVANCE_PROMPT = (
    "Look at this HKEX Application Proof Multi-Files index page (the "
    "company name + chapter list). Determine whether this issuer's CURRENT "
    "main business is in life sciences. Watch for rebranded entities -- if "
    "the English name says e.g. 'XYZ Food Group (formerly known as XYZ "
    "Biotechnology Group)', the CURRENT business is FOOD, not biotech, so "
    "is_lifesci=false and rebrand_concern=true. Be conservative: only return "
    "is_lifesci=true when life sciences is unambiguously the main business."
)


def apply_llm_relevance(
    candidate: Dict[str, Any],
    relevance_data: Dict[str, Any],
) -> Dict[str, Any]:
    """Merge LLM relevance check result into a candidate dict.

    Mutates candidate in place. Adds:
      - candidate['llm_relevance'] = relevance_data (the raw LLM dict)
      - candidate['qc_flags'] += relevant tags
      - if rebrand_concern=True or is_lifesci=False -> sets
        candidate['llm_downgrade_reason'] for the Phase 0 Gate 1 renderer
        to surface in the LLM DOWNGRADED bucket.
    """
    candidate.setdefault("qc_flags", [])
    candidate["llm_relevance"] = relevance_data or {}
    if not isinstance(relevance_data, dict):
        return candidate
    is_ls = relevance_data.get("is_lifesci")
    rebrand = relevance_data.get("rebrand_concern")
    current = relevance_data.get("current_business", "")[:120]
    if is_ls is False or rebrand is True:
        reason = "rebranded out of LS" if rebrand else "LLM says not life-sciences"
        candidate["llm_downgrade_reason"] = "{}: {}".format(reason, current)
        candidate["qc_flags"].append("llm_downgrade_candidate")
    return candidate


# ============================================================================
# Extraction Health Dashboard (#7) -- aggregate extraction success per col
# ============================================================================

def compute_extraction_health(staging_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Aggregate per-column extraction success across staging rows.

    Returns a dict suitable for rendering in the Phase 4 Gate 2 diff:
      {
        "n_rows": int,
        "by_col": {
          "F": {"success": int, "low_conf": int, "blank": int, "rate": float},
          ...
        },
        "abort_recommended": bool,        # any col < 50% success
        "block_cols": List[str],          # cols below 50% threshold
      }

    A cell is "success" if row_draft[col] is non-empty AND confidence is not
    "low" AND no firecrawl_pending flag exists for that col.
    """
    cols = ("F", "G", "H", "I", "J", "K", "L", "M")
    out: Dict[str, Any] = {"n_rows": len(staging_rows), "by_col": {}}
    if not staging_rows:
        out["abort_recommended"] = False
        out["block_cols"] = []
        return out
    block_cols: List[str] = []
    for col in cols:
        success = 0
        low_conf = 0
        blank = 0
        for row in staging_rows:
            rd = row.get("row_draft", {})
            conf = (row.get("_confidence") or {}).get(col, "low")
            qc = row.get("_qc_flags", [])
            v = rd.get(col)
            pending_flag = "firecrawl_pending_col_{}".format(col)
            is_pending = pending_flag in qc
            if v is None or v == "" or is_pending:
                blank += 1
                continue
            if conf == "low":
                low_conf += 1
                continue
            success += 1
        n = len(staging_rows)
        rate = success / n if n else 0.0
        out["by_col"][col] = {
            "success": success,
            "low_conf": low_conf,
            "blank": blank,
            "rate": round(rate, 3),
        }
        if rate < 0.5:
            block_cols.append(col)
    out["block_cols"] = block_cols
    out["abort_recommended"] = bool(block_cols)
    return out


def render_extraction_health(health: Dict[str, Any]) -> str:
    """Format compute_extraction_health output for Gate 2 console output."""
    lines = ["=== Extraction Health (NEW + REFRESH being written) ==="]
    n = health.get("n_rows", 0)
    if n == 0:
        return "\n".join(lines + ["  (no rows)"])
    col_labels = {
        "F": "F (structure)", "G": "G (business)", "H": "H (sector)",
        "I": "I (sponsor)", "J": "J (revenue)", "K": "K (NI)",
        "L": "L (cash)", "M": "M (lead asset)",
    }
    for col, info in health.get("by_col", {}).items():
        s = info["success"]
        lc = info["low_conf"]
        b = info["blank"]
        rate = info["rate"]
        if rate >= 0.9:
            mark = "OK "
        elif rate >= 0.5:
            mark = "WARN"
        else:
            mark = "FAIL"
        suffix = ""
        if lc > 0:
            suffix += "  ({} low-conf)".format(lc)
        if b > 0:
            suffix += "  ({} blank)".format(b)
        lines.append(
            "  {:6s} {}/{}  {:3.0f}%  {}{}".format(
                mark, s, n, rate * 100, col_labels.get(col, col), suffix
            )
        )
    if health.get("abort_recommended"):
        lines.append("")
        lines.append("  >> ABORT RECOMMENDED: cols {} below 50% success.".format(
            ", ".join(health["block_cols"])
        ))
        lines.append("     Investigate root cause before Phase 5 write.")
    return "\n".join(lines)


# ============================================================================
# Col B status lifecycle (#5) -- compute "Expiring in N month" / "Expired"
# ============================================================================

def compute_status(
    filing_date: Any,
    today: Optional[datetime.date] = None,
    validity_months: int = 6,
) -> Optional[str]:
    """Compute col B status based on filing date and HKEX 6-month validity.

    HKEX Application Proof has a 6-month validity window. Status semantics:
      - 0-3 months since filing: None (fresh, no marker)
      - 4-5 months since filing: 'Expiring in 2 month' / 'Expiring in 1 month'
      - >= 6 months: 'Expired' (caller should also append † to col D for Expired
                     rows, but this function returns only the status text)
    Returns None when filing_date is unparseable.

    Used during REFRESH (where the filing date just changed -> status resets
    to None) and during a periodic refresh of all rows in the master.
    """
    if filing_date is None:
        return None
    if isinstance(filing_date, datetime.datetime):
        fd = filing_date.date()
    elif isinstance(filing_date, datetime.date):
        fd = filing_date
    else:
        try:
            fd = datetime.datetime.strptime(
                str(filing_date)[:10], "%Y-%m-%d"
            ).date()
        except ValueError:
            return None
    if today is None:
        today = datetime.date.today()
    days_since = (today - fd).days
    months_since = days_since // 30
    months_remaining = validity_months - months_since
    if months_remaining <= 0:
        return "Expired"
    if months_remaining <= 2:
        return "Expiring in {} month".format(months_remaining)
    return None


# ============================================================================
# Col F enum schema check (#6) -- validate before Phase 5 write
# ============================================================================

VALID_COL_F_VALUES = ("H-share", "Red Chip", "VIE")


def validate_col_f_value(value: Any) -> Tuple[bool, Optional[str]]:
    """Validate a col F value against the post-May-2026 3-option enum.

    Returns:
      (is_valid: bool, suggested_replacement: Optional[str])
      - ('H-share', None)               -> valid, no migration needed
      - (False, 'Red Chip')              -> 'Cayman holdco' / 'BVI holdco' map to 'Red Chip'
      - (False, None)                    -> unknown value, manual review needed

    Used by Phase 5 pre-write QC AND by scripts/migrate_col_f_v2.py.
    """
    if value is None or value == "":
        return (False, None)
    s = str(value).strip()
    if s in VALID_COL_F_VALUES:
        return (True, None)
    # Legacy v1 enum -> migration mapping
    if s in ("Cayman holdco", "BVI holdco", "Bermuda holdco"):
        return (False, "Red Chip")
    # Common typos / casing
    s_lower = s.lower().replace("-", "").replace(" ", "")
    if s_lower in ("hshare", "hshares"):
        return (False, "H-share")
    if s_lower in ("redchip", "redchips"):
        return (False, "Red Chip")
    if s_lower == "vie":
        return (False, "VIE")
    return (False, None)

