"""migrate_col_f_v2.py -- one-shot migration from 5-option F enum to 3-option.

Rewrites col F values in `tracker/a1_pipeline_tracker.xlsx`:
  - 'Cayman holdco' -> 'Red Chip' (with N-column footnote: '[Cayman-incorporated]')
  - 'BVI holdco'    -> 'Red Chip' (with N-column footnote: '[BVI-incorporated]')
  - 'Bermuda holdco' -> 'Red Chip' (with N-column footnote: '[Bermuda-incorporated]')
  - typos / casing  -> canonical via validate_col_f_value()

Also reports rows where F is blank (legacy gap) for the user to backfill in
a follow-up pass. Does NOT modify the live tracker until --apply is passed;
the default is dry-run with diff preview.

Usage:
    python3 scripts/migrate_col_f_v2.py            # dry run
    python3 scripts/migrate_col_f_v2.py --apply    # write back
"""
import argparse
import sys
import os
from copy import copy

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import hkex_scraper as hs

LIVE_PATH = "tracker/a1_pipeline_tracker.xlsx"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write back; default is dry-run")
    parser.add_argument("--path", default=LIVE_PATH, help="Tracker path")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print("ERROR: tracker not found at {}".format(args.path))
        return 1

    wb = openpyxl.load_workbook(args.path)
    ws = wb.active

    # Find last data row
    last_data = ws.max_row
    while last_data > 2 and not ws.cell(row=last_data, column=4).value:
        last_data -= 1
    print("Tracker: {} ({} data rows)".format(args.path, last_data - 2))
    print()

    migrations = []   # (row, old, new, suffix_for_N)
    blanks = []       # rows where F is None / ''
    invalid = []      # rows where F is unknown value with no migration mapping

    for r in range(3, last_data + 1):
        old_f = ws.cell(row=r, column=6).value
        company = ws.cell(row=r, column=4).value or "?"
        if old_f is None or (isinstance(old_f, str) and not old_f.strip()):
            blanks.append((r, company))
            continue
        valid, replacement = hs.validate_col_f_value(old_f)
        if valid:
            continue
        if replacement is None:
            invalid.append((r, company, old_f))
            continue
        # Determine N suffix for legacy Cayman/BVI/Bermuda holdco values
        suffix = None
        old_str = str(old_f).strip()
        if old_str == "Cayman holdco":
            suffix = "[Cayman-incorporated]"
        elif old_str == "BVI holdco":
            suffix = "[BVI-incorporated]"
        elif old_str == "Bermuda holdco":
            suffix = "[Bermuda-incorporated]"
        migrations.append((r, company, old_f, replacement, suffix))

    print("=" * 70)
    print("MIGRATION PLAN (col F enum: 5 -> 3 values)")
    print("=" * 70)

    if migrations:
        print("\n>> {} rows to MIGRATE:".format(len(migrations)))
        for r, co, old, new, suf in migrations:
            note = " + col N suffix '{}'".format(suf) if suf else ""
            print("   r{:3d}  '{}' -> '{}'{}".format(r, old, new, note))
            print("         {} ({})".format(str(co)[:60],
                  ws.cell(row=r, column=5).value or ""))
    else:
        print("\n>> No legacy values to migrate. (All F values are H-share / Red Chip / VIE.)")

    if blanks:
        print("\n>> {} rows have BLANK col F (legacy gap, not auto-fixable):".format(len(blanks)))
        for r, co in blanks[:15]:
            print("   r{:3d}  {}".format(r, str(co)[:70]))
        if len(blanks) > 15:
            print("   ... ({} more)".format(len(blanks) - 15))
        print("   These need manual backfill via /a1-pipeline-update REFRESH or direct edit.")

    if invalid:
        print("\n>> {} rows have UNKNOWN col F values (manual review):".format(len(invalid)))
        for r, co, old in invalid:
            print("   r{:3d}  '{}'  ({})".format(r, old, str(co)[:60]))

    if not args.apply:
        print("\n[DRY RUN] No changes written. Re-run with --apply to commit.")
        return 0

    if not migrations:
        print("\nNothing to migrate. Exiting without write.")
        return 0

    # Apply migrations: F update + N suffix append
    for r, co, old, new, suf in migrations:
        f_cell = ws.cell(row=r, column=6)
        f_cell.value = new
        # Preserve font (skill enforces Arial 8 black; keep that)
        # Append suffix to N if present and not already there
        if suf:
            n_cell = ws.cell(row=r, column=14)
            existing_n = n_cell.value or ""
            if suf not in existing_n:
                if existing_n:
                    n_cell.value = (existing_n.rstrip() + " " + suf).strip()
                else:
                    n_cell.value = suf

    # Write back (with snapshot first)
    snapshot_path = args.path.replace(
        ".xlsx", "_pre_col_f_v2_migration_backup.xlsx"
    )
    if not os.path.exists(snapshot_path):
        import shutil
        shutil.copy2(args.path, snapshot_path)
        print("\nBackup saved -> {}".format(snapshot_path))

    wb.save(args.path)
    print("Wrote {} migrations -> {}".format(len(migrations), args.path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
